from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from logging import getLogger

logger = getLogger(__name__)


class ExcelSheet:
    def __init__(self, filename: str, min_row: int = 2):
        """
        Класс, представляющий лист Excel.

        Args:
            filename (str): Имя файла Excel.
        """
        self.filename = filename
        self.min_row = min_row
        self.__sheet_data = []
        self.load_workbook()

    def load_workbook(self) -> None:
        """
        Загружает рабочую книгу и устанавливает активный лист.
        """
        try:
            self.wb = load_workbook(filename=self.filename)
            self.sheet = self.wb.active
            self.load_data()
        except FileNotFoundError:
            logger.error("Файл не найден: %s", self.filename)
        except PermissionError:
            logger.error("Ошибка доступа: %s", self.filename)

    def load_data(self) -> None:
        """
        Загружает данные из файла Excel.
        """

        self.__sheet_data = [row for row in self.sheet.iter_rows(min_row=self.min_row, values_only=True)]

    def get_data(self) -> List:
        """
        Возвращает данные листа Excel.

        Returns:
            List: Данные листа Excel.
        """
        logger.debug("Exel-таблица считана")
        return self.__sheet_data

    def highlight_cells(self, cells_to_highlight: Dict[int, List[int]]) -> None:
        """
        Закрашивает указанные ячейки в красный цвет.

        Args:
            cells_to_highlight (Dict[int, List[int]]): Словарь, где ключ - номер строки (начиная с 1),
                                                       значение - список номеров столбцов (начиная с 1).
        """
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for row in range(self.min_row, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column - 1):
                cell = self.sheet.cell(row=row, column=col)
                if row in cells_to_highlight and col in cells_to_highlight[row]:
                    cell.fill = red_fill
                else:
                    cell.fill = white_fill

        # Сохранение изменений в файл
        try:
            self.wb.save(self.filename)
        except PermissionError:
            logger.error("Ошибка доступа: %s", self.filename)
