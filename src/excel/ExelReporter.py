import os
from typing import List
from datetime import datetime
from openpyxl import Workbook, load_workbook
from logging import getLogger

logger = getLogger(__name__)


class Reporter:
    def __init__(self, output_directory: str):
        self.output_directory = output_directory
        self.filename = os.path.join(self.output_directory,
                                     f"Отчет_{datetime.today().strftime('%d.%m.%y')}.xlsx")

    def generate_report(self, reporter_list: List[List]) -> None:
        """
        Создаёт отчёт в формате Excel на основе предоставленных данных.

        :param reporter_list: Список с данными для отчёта
        """
        try:
            if os.path.exists(self.filename):
                wb = load_workbook(self.filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет"
                headers = [
                    "Номер задачи в JIRA", "Название процесса", "Аналитик",
                    "Путь к папке, которую нужно очищать", "Наименование удаленных папок/ файлов",
                    "Дата/ Время начала обработки", "Дата / Время завершения обработки", "Статус", "Комментарий"
                ]
                ws.append(headers)

            for report in reporter_list:
                task_number, process_name, analyst, folder_path, report_dict, start_time, end_time = report

                if "Нет файлов на удаление" in report_dict:
                    ws.append([
                        task_number, process_name, analyst, folder_path, "Нет файлов на удаление", start_time, end_time,
                        report_dict['Нет файлов на удаление'].status, report_dict['Нет файлов на удаление'].comment
                    ])
                    continue

                else:
                    ws.append([
                        task_number, process_name, analyst, folder_path,
                        f"Список файлов на удаление ({len(report_dict)} эл)", start_time, end_time, "", ""
                    ])
                for path, result in report_dict.items():
                    ws.append([
                        "", "", "", "", path,
                        "", "", result.status, result.comment
                    ])
            wb.save(self.filename)
            logger.info("Отчёт сохранён по пути: %s", self.filename)
        except PermissionError:
            logger.error("Не удалось сохранить отчёт. Возможно, файл открыт в другой программе.")
        except Exception as e:
            logger.error("Произошла ошибка при генерации отчёта: %s", e)
